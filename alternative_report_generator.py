import os
import sys
import json
from datetime import datetime
import sqlite3
import base64
from io import BytesIO

# For PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# For Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Data model classes (same as original)
class Project:
    def __init__(self, id, project_id_str, name):
        self.id = id
        self.project_id_str = project_id_str
        self.name = name

class SubJob:
    def __init__(self, id, project_id, sub_job_id_str, name):
        self.id = id
        self.project_id = project_id
        self.sub_job_id_str = sub_job_id_str
        self.name = name

class CostCode:
    def __init__(self, id, project_id, cost_code_id_str, description, discipline, rule_of_credit_id):
        self.id = id
        self.project_id = project_id
        self.cost_code_id_str = cost_code_id_str
        self.description = description
        self.discipline = discipline
        self.rule_of_credit_id = rule_of_credit_id

class RuleOfCredit:
    def __init__(self, id, name, steps_json):
        self.id = id
        self.name = name
        self.steps_json = steps_json
        try:
            self.steps = json.loads(steps_json) if steps_json else []
        except json.JSONDecodeError:
            self.steps = []

class WorkItem:
    def __init__(self, id, project_id, sub_job_id, cost_code_id, work_item_id_str, description, 
                 budgeted_quantity, earned_quantity, percent_complete_quantity, 
                 budgeted_man_hours, earned_man_hours, percent_complete_hours, 
                 unit_of_measure, progress_json):
        self.id = id
        self.project_id = project_id
        self.sub_job_id = sub_job_id
        self.cost_code_id = cost_code_id
        self.work_item_id_str = work_item_id_str
        self.description = description
        self.budgeted_quantity = budgeted_quantity
        self.earned_quantity = earned_quantity
        self.percent_complete_quantity = percent_complete_quantity
        self.budgeted_man_hours = budgeted_man_hours
        self.earned_man_hours = earned_man_hours
        self.percent_complete_hours = percent_complete_hours
        self.unit_of_measure = unit_of_measure
        self.progress_json = progress_json
        self.roc_steps_progress = {}
        if self.progress_json:
            try:
                progress_data = json.loads(self.progress_json)
                # Ensure progress_data is a list of dictionaries with "step_name" and "current_complete_percentage"
                if isinstance(progress_data, list):
                    for step_progress in progress_data:
                        if isinstance(step_progress, dict) and "step_name" in step_progress and "current_complete_percentage" in step_progress:
                            self.roc_steps_progress[step_progress["step_name"]] = float(step_progress["current_complete_percentage"])
                        elif isinstance(step_progress, dict) and "name" in step_progress and "percentage" in step_progress: # Legacy format
                             self.roc_steps_progress[step_progress["name"]] = float(step_progress["percentage"])
                elif isinstance(progress_data, dict): # Older format where keys are step names
                    for step_name, percentage in progress_data.items():
                        self.roc_steps_progress[step_name] = float(percentage)

            except (json.JSONDecodeError, TypeError, ValueError) as e:
                print(f"Error parsing progress data for work item {self.id}: {e}")

class AlternativeReportGenerator:
    """
    A class to generate PDF and Excel reports for the Magellan EV Tracker using ReportLab and openpyxl.
    This class handles both project-wide and sub-job specific reports.
    """
    
    def __init__(self, db_path, logo_path):
        self.db_path = db_path
        self.logo_path = logo_path
        self.conn = None
        self.logo_data = None
        self._connect_db()
        self._load_logo()
    
    def _connect_db(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
    
    def _load_logo(self):
        try:
            with open(self.logo_path, "rb") as logo_file:
                self.logo_data = logo_file.read()
        except Exception as e:
            print(f"Warning: Could not load logo from {self.logo_path}: {e}")
            self.logo_data = None
    
    def _get_report_data(self, project_id, sub_job_id=None):
        """Get all data needed for report generation"""
        cursor = self.conn.cursor()
        
        # Get project info
        cursor.execute("SELECT id, project_id_str, name FROM project WHERE id = ?", (project_id,))
        project_row = cursor.fetchone()
        if not project_row:
            raise ValueError(f"Project with ID {project_id} not found")
        
        project = Project(project_row['id'], project_row['project_id_str'], project_row['name'])
        
        # Get sub-job info if specified
        sub_job = None
        if sub_job_id:
            cursor.execute("SELECT id, project_id, sub_job_id_str, name FROM sub_job WHERE id = ?", (sub_job_id,))
            sub_job_row = cursor.fetchone()
            if not sub_job_row:
                raise ValueError(f"Sub-job with ID {sub_job_id} not found")
            
            sub_job = SubJob(sub_job_row['id'], sub_job_row['project_id'], sub_job_row['sub_job_id_str'], sub_job_row['name'])
        
        # Get work items
        work_items_query = """
        SELECT 
            wi.id, wi.project_id, wi.sub_job_id, wi.cost_code_id, wi.work_item_id_str, wi.description,
            wi.budgeted_quantity, wi.earned_quantity, wi.percent_complete_quantity,
            wi.budgeted_man_hours, wi.earned_man_hours, wi.percent_complete_hours,
            wi.unit_of_measure, wi.progress_json,
            cc.cost_code_id_str, cc.description as cc_description, cc.discipline, cc.rule_of_credit_id,
            roc.name as roc_name, roc.steps_json
        FROM work_item wi
        JOIN cost_code cc ON wi.cost_code_id = cc.id
        LEFT JOIN rule_of_credit roc ON cc.rule_of_credit_id = roc.id
        WHERE wi.project_id = ?
        """
        
        params = [project_id]
        if sub_job_id:
            work_items_query += " AND wi.sub_job_id = ?"
            params.append(sub_job_id)
        
        cursor.execute(work_items_query, params)
        work_item_rows = cursor.fetchall()
        
        # Process work items and organize by discipline and cost code
        all_work_items = []
        grouped_data = {}
        project_totals = {
            'budgeted_hours': 0,
            'earned_hours': 0,
            'percent_complete_hours': 0,
            'budgeted_quantity': 0,
            'earned_quantity': 0,
            'percent_complete_quantity': 0
        }
        
        all_roc_step_names = set()
        
        for row in work_item_rows:
            work_item = WorkItem(
                row['id'], row['project_id'], row['sub_job_id'], row['cost_code_id'],
                row['work_item_id_str'], row['description'],
                row['budgeted_quantity'], row['earned_quantity'], row['percent_complete_quantity'],
                row['budgeted_man_hours'], row['earned_man_hours'], row['percent_complete_hours'],
                row['unit_of_measure'], row['progress_json']
            )
            
            all_work_items.append(work_item)
            
            # Get discipline and cost code
            discipline = row['discipline'] or "Unassigned"
            cost_code = row['cost_code_id_str']
            
            # Initialize discipline if not exists
            if discipline not in grouped_data:
                grouped_data[discipline] = {
                    'subtotals': {
                        'budgeted_hours': 0,
                        'earned_hours': 0,
                        'percent_complete_hours': 0,
                        'budgeted_quantity': 0,
                        'earned_quantity': 0,
                        'percent_complete_quantity': 0
                    },
                    'cost_codes': {}
                }
            
            # Initialize cost code if not exists
            if cost_code not in grouped_data[discipline]['cost_codes']:
                grouped_data[discipline]['cost_codes'][cost_code] = {
                    'subtotals': {
                        'budgeted_hours': 0,
                        'earned_hours': 0,
                        'percent_complete_hours': 0,
                        'budgeted_quantity': 0,
                        'earned_quantity': 0,
                        'percent_complete_quantity': 0
                    },
                    'work_items': [],
                    'roc_step_headers': []
                }
                
                # Parse RoC steps for this cost code
                if row['steps_json']:
                    try:
                        roc_data = json.loads(row['steps_json'])
                        if isinstance(roc_data, dict) and 'steps' in roc_data:
                            step_headers = []
                            for step in roc_data['steps']:
                                if isinstance(step, dict) and 'name' in step:
                                    step_name = step['name']
                                    step_headers.append(step_name)
                                    all_roc_step_names.add(step_name)
                            
                            # Ensure we have exactly 7 columns (pad with empty strings if needed)
                            while len(step_headers) < 7:
                                step_headers.append("")
                            
                            # Truncate if more than 7
                            if len(step_headers) > 7:
                                step_headers = step_headers[:7]
                                
                            grouped_data[discipline]['cost_codes'][cost_code]['roc_step_headers'] = step_headers
                    except (json.JSONDecodeError, TypeError) as e:
                        print(f"Error parsing RoC steps for cost code {cost_code}: {e}")
                        grouped_data[discipline]['cost_codes'][cost_code]['roc_step_headers'] = [""] * 7
                else:
                    grouped_data[discipline]['cost_codes'][cost_code]['roc_step_headers'] = [""] * 7
            
            # Add work item to cost code
            grouped_data[discipline]['cost_codes'][cost_code]['work_items'].append(work_item)
            
            # Update subtotals
            cc_subtotals = grouped_data[discipline]['cost_codes'][cost_code]['subtotals']
            disc_subtotals = grouped_data[discipline]['subtotals']
            
            for subtotals in [cc_subtotals, disc_subtotals, project_totals]:
                subtotals['budgeted_hours'] += work_item.budgeted_man_hours or 0
                subtotals['earned_hours'] += work_item.earned_man_hours or 0
                subtotals['budgeted_quantity'] += work_item.budgeted_quantity or 0
                subtotals['earned_quantity'] += work_item.earned_quantity or 0
        
        # Calculate percentages
        for discipline, discipline_data in grouped_data.items():
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                subtotals = cost_code_data['subtotals']
                if subtotals['budgeted_hours'] > 0:
                    subtotals['percent_complete_hours'] = (subtotals['earned_hours'] / subtotals['budgeted_hours']) * 100
                if subtotals['budgeted_quantity'] > 0:
                    subtotals['percent_complete_quantity'] = (subtotals['earned_quantity'] / subtotals['budgeted_quantity']) * 100
            
            subtotals = discipline_data['subtotals']
            if subtotals['budgeted_hours'] > 0:
                subtotals['percent_complete_hours'] = (subtotals['earned_hours'] / subtotals['budgeted_hours']) * 100
            if subtotals['budgeted_quantity'] > 0:
                subtotals['percent_complete_quantity'] = (subtotals['earned_quantity'] / subtotals['budgeted_quantity']) * 100
        
        if project_totals['budgeted_hours'] > 0:
            project_totals['percent_complete_hours'] = (project_totals['earned_hours'] / project_totals['budgeted_hours']) * 100
        if project_totals['budgeted_quantity'] > 0:
            project_totals['percent_complete_quantity'] = (project_totals['earned_quantity'] / project_totals['budgeted_quantity']) * 100
        
        # Sort data
        sorted_grouped_data = {}
        for discipline in sorted(grouped_data.keys()):
            sorted_grouped_data[discipline] = {
                'subtotals': grouped_data[discipline]['subtotals'],
                'cost_codes': {}
            }
            
            for cost_code in sorted(grouped_data[discipline]['cost_codes'].keys()):
                sorted_grouped_data[discipline]['cost_codes'][cost_code] = grouped_data[discipline]['cost_codes'][cost_code]
                sorted_grouped_data[discipline]['cost_codes'][cost_code]['work_items'].sort(key=lambda wi: wi.work_item_id_str or "")
        
        return {
            'project': project,
            'sub_job': sub_job,
            'all_work_items': all_work_items,
            'grouped_work_items': sorted_grouped_data,
            'project_totals': project_totals,
            'all_roc_step_names': sorted(list(all_roc_step_names)),
            'report_date': datetime.now().strftime("%Y-%m-%d")
        }
    
    def generate_hours_pdf(self, project_id, output_path, sub_job_id=None):
        """Generate hours PDF report using ReportLab"""
        data = self._get_report_data(project_id, sub_job_id)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=14,
            alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER
        )
        project_title_style = ParagraphStyle(
            'ProjectTitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        project_id_style = ParagraphStyle(
            'ProjectID',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER
        )
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=8
        )
        right_align_style = ParagraphStyle(
            'RightAlign',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_RIGHT
        )
        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=6
        )
        
        # Elements to add to PDF
        elements = []
        
        # Header table
        if self.logo_data:
            logo = Image(BytesIO(self.logo_data))
            logo.drawHeight = 0.5*inch
            logo.drawWidth = 1*inch
        else:
            logo = Paragraph("LOGO", normal_style)
        
        # Header data - Removed space between title and subtitle
        header_data = [
            [
                logo,
                [
                    Paragraph("PROGRESS DETAIL<br/>Sub Job by Hours" if data['sub_job'] else "PROGRESS DETAIL<br/>Total Project by Hours", title_style),
                    Paragraph(data['project'].name, project_title_style),
                    Paragraph(data['project'].project_id_str, project_id_style)
                ],
                [
                    Paragraph(f"Page: 1", right_align_style),
                    Paragraph("Progress Thru:", right_align_style),
                    Paragraph(data['report_date'], right_align_style)
                ]
            ]
        ]
        
        # Header table style
        header_table_style = TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('SPAN', (0, 0), (0, 0)),
            ('SPAN', (1, 0), (1, 0)),
            ('SPAN', (2, 0), (2, 0)),
        ])
        
        header_table = Table(header_data, colWidths=[2*inch, 5*inch, 2*inch])
        header_table.setStyle(header_table_style)
        elements.append(header_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # If sub-job, add sub-job title
        if data['sub_job']:
            elements.append(Paragraph(f"{data['sub_job'].sub_job_id_str} - {data['sub_job'].name}", styles['Heading3']))
            elements.append(Spacer(1, 0.1*inch))
        
        # Main report table
        # Column headers
        table_data = [
            [
                Paragraph("Work Item", header_style),
                Paragraph("Description", header_style),
                Paragraph("Budgeted Hours", header_style),
                Paragraph("Earned Hours", header_style),
                Paragraph("% Complete", header_style),
                Paragraph("Rules of Credit Steps", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style)
            ]
        ]
        
        # Merge the Rules of Credit Steps cells
        table_style = [
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 0), (-1, -1), 'CENTER'),
            ('SPAN', (5, 0), (-1, 0)),  # Merge Rules of Credit Steps header
            ('FONTSIZE', (0, 0), (-1, 0), 8),
        ]
        
        # Add data rows
        row_index = 1
        for discipline, discipline_data in data['grouped_work_items'].items():
            # Discipline header
            table_data.append([Paragraph(discipline, header_style)] + [""] * 11)
            table_style.append(('SPAN', (0, row_index), (-1, row_index)))
            table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey))
            table_style.append(('ALIGN', (0, row_index), (0, row_index), 'LEFT'))
            row_index += 1
            
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                # Cost code header with RoC step names
                cc_row = [Paragraph(cost_code, header_style), "", "", "", ""]
                
                # Add RoC step headers
                for step_name in cost_code_data['roc_step_headers']:
                    if step_name:
                        cc_row.append(Paragraph(step_name, small_style))
                    else:
                        cc_row.append("")
                
                table_data.append(cc_row)
                table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey))
                table_style.append(('SPAN', (0, row_index), (4, row_index)))
                table_style.append(('FONTSIZE', (5, row_index), (-1, row_index), 6))  # Smaller font for RoC headers
                row_index += 1
                
                # Work items
                for work_item in cost_code_data['work_items']:
                    wi_row = [
                        Paragraph(work_item.work_item_id_str, normal_style),
                        Paragraph(work_item.description, normal_style),
                        Paragraph(f"{work_item.budgeted_man_hours:.2f}", normal_style),
                        Paragraph(f"{work_item.earned_man_hours:.2f}", normal_style),
                        Paragraph(f"{work_item.percent_complete_hours:.1f}%", normal_style)
                    ]
                    
                    # Add RoC step progress
                    for step_name in cost_code_data['roc_step_headers']:
                        if step_name and step_name in work_item.roc_steps_progress:
                            wi_row.append(Paragraph(f"{work_item.roc_steps_progress[step_name]:.1f}%", small_style))
                        elif step_name:
                            wi_row.append(Paragraph("0.0%", small_style))
                        else:
                            wi_row.append("")
                    
                    table_data.append(wi_row)
                    table_style.append(('FONTSIZE', (5, row_index), (-1, row_index), 6))  # Smaller font for RoC values
                    row_index += 1
                
                # Cost code subtotal
                cc_subtotal_row = [
                    Paragraph("Cost Code Subtotal", normal_style),
                    "",
                    Paragraph(f"{cost_code_data['subtotals']['budgeted_hours']:.2f}", normal_style),
                    Paragraph(f"{cost_code_data['subtotals']['earned_hours']:.2f}", normal_style),
                    Paragraph(f"{cost_code_data['subtotals']['percent_complete_hours']:.1f}%", normal_style),
                    "", "", "", "", "", "", ""
                ]
                table_data.append(cc_subtotal_row)
                table_style.append(('SPAN', (0, row_index), (1, row_index)))
                table_style.append(('SPAN', (5, row_index), (-1, row_index)))
                table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
                row_index += 1
            
            # Discipline subtotal
            disc_subtotal_row = [
                Paragraph("Discipline Subtotal", normal_style),
                "",
                Paragraph(f"{discipline_data['subtotals']['budgeted_hours']:.2f}", normal_style),
                Paragraph(f"{discipline_data['subtotals']['earned_hours']:.2f}", normal_style),
                Paragraph(f"{discipline_data['subtotals']['percent_complete_hours']:.1f}%", normal_style),
                "", "", "", "", "", "", ""
            ]
            table_data.append(disc_subtotal_row)
            table_style.append(('SPAN', (0, row_index), (1, row_index)))
            table_style.append(('SPAN', (5, row_index), (-1, row_index)))
            table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
            row_index += 1
        
        # Total row
        total_row = [
            Paragraph("Total", normal_style),
            "",
            Paragraph(f"{data['project_totals']['budgeted_hours']:.2f}", normal_style),
            Paragraph(f"{data['project_totals']['earned_hours']:.2f}", normal_style),
            Paragraph(f"{data['project_totals']['percent_complete_hours']:.1f}%", normal_style),
            "", "", "", "", "", "", ""
        ]
        table_data.append(total_row)
        table_style.append(('SPAN', (0, row_index), (1, row_index)))
        table_style.append(('SPAN', (5, row_index), (-1, row_index)))
        table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
        
        # Create table with column widths
        col_widths = [
            0.8*inch,  # Work Item (10%)
            1.6*inch,  # Description (20%)
            0.8*inch,  # Budgeted Hours
            0.8*inch,  # Earned Hours
            0.8*inch,  # % Complete
            0.6*inch,  # RoC 1
            0.6*inch,  # RoC 2
            0.6*inch,  # RoC 3
            0.6*inch,  # RoC 4
            0.6*inch,  # RoC 5
            0.6*inch,  # RoC 6
            0.6*inch   # RoC 7
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return output_path
    
    def generate_quantities_pdf(self, project_id, output_path, sub_job_id=None):
        """Generate quantities PDF report using ReportLab"""
        data = self._get_report_data(project_id, sub_job_id)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=14,
            alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER
        )
        project_title_style = ParagraphStyle(
            'ProjectTitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        project_id_style = ParagraphStyle(
            'ProjectID',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER
        )
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=8
        )
        right_align_style = ParagraphStyle(
            'RightAlign',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_RIGHT
        )
        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=6
        )
        
        # Elements to add to PDF
        elements = []
        
        # Header table
        if self.logo_data:
            logo = Image(BytesIO(self.logo_data))
            logo.drawHeight = 0.5*inch
            logo.drawWidth = 1*inch
        else:
            logo = Paragraph("LOGO", normal_style)
        
        # Header data - Removed space between title and subtitle
        header_data = [
            [
                logo,
                [
                    Paragraph("QUANTITY DETAIL<br/>Sub Job by Quantities" if data['sub_job'] else "QUANTITY DETAIL<br/>Total Project by Quantities", title_style),
                    Paragraph(data['project'].name, project_title_style),
                    Paragraph(data['project'].project_id_str, project_id_style)
                ],
                [
                    Paragraph(f"Page: 1", right_align_style),
                    Paragraph("Progress Thru:", right_align_style),
                    Paragraph(data['report_date'], right_align_style)
                ]
            ]
        ]
        
        # Header table style
        header_table_style = TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('SPAN', (0, 0), (0, 0)),
            ('SPAN', (1, 0), (1, 0)),
            ('SPAN', (2, 0), (2, 0)),
        ])
        
        header_table = Table(header_data, colWidths=[2*inch, 5*inch, 2*inch])
        header_table.setStyle(header_table_style)
        elements.append(header_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # If sub-job, add sub-job title
        if data['sub_job']:
            elements.append(Paragraph(f"{data['sub_job'].sub_job_id_str} - {data['sub_job'].name}", styles['Heading3']))
            elements.append(Spacer(1, 0.1*inch))
        
        # Main report table
        # Column headers
        table_data = [
            [
                Paragraph("Work Item", header_style),
                Paragraph("Description", header_style),
                Paragraph("UOM", header_style),
                Paragraph("Budgeted Qty", header_style),
                Paragraph("Earned Qty", header_style),
                Paragraph("% Complete", header_style),
                Paragraph("Rules of Credit Steps", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style),
                Paragraph("", header_style)
            ]
        ]
        
        # Merge the Rules of Credit Steps cells
        table_style = [
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (5, -1), 'RIGHT'),
            ('ALIGN', (6, 0), (-1, -1), 'CENTER'),
            ('SPAN', (6, 0), (-1, 0)),  # Merge Rules of Credit Steps header
            ('FONTSIZE', (0, 0), (-1, 0), 8),
        ]
        
        # Add data rows
        row_index = 1
        for discipline, discipline_data in data['grouped_work_items'].items():
            # Discipline header
            table_data.append([Paragraph(discipline, header_style)] + [""] * 12)
            table_style.append(('SPAN', (0, row_index), (-1, row_index)))
            table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey))
            table_style.append(('ALIGN', (0, row_index), (0, row_index), 'LEFT'))
            row_index += 1
            
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                # Cost code header with RoC step names
                cc_row = [Paragraph(cost_code, header_style), "", "", "", "", ""]
                
                # Add RoC step headers
                for step_name in cost_code_data['roc_step_headers']:
                    if step_name:
                        cc_row.append(Paragraph(step_name, small_style))
                    else:
                        cc_row.append("")
                
                table_data.append(cc_row)
                table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey))
                table_style.append(('SPAN', (0, row_index), (5, row_index)))
                table_style.append(('FONTSIZE', (6, row_index), (-1, row_index), 6))  # Smaller font for RoC headers
                row_index += 1
                
                # Work items
                for work_item in cost_code_data['work_items']:
                    wi_row = [
                        Paragraph(work_item.work_item_id_str, normal_style),
                        Paragraph(work_item.description, normal_style),
                        Paragraph(work_item.unit_of_measure or "", normal_style),
                        Paragraph(f"{work_item.budgeted_quantity:.2f}", normal_style),
                        Paragraph(f"{work_item.earned_quantity:.2f}", normal_style),
                        Paragraph(f"{work_item.percent_complete_quantity:.1f}%", normal_style)
                    ]
                    
                    # Add RoC step progress
                    for step_name in cost_code_data['roc_step_headers']:
                        if step_name and step_name in work_item.roc_steps_progress:
                            wi_row.append(Paragraph(f"{work_item.roc_steps_progress[step_name]:.1f}%", small_style))
                        elif step_name:
                            wi_row.append(Paragraph("0.0%", small_style))
                        else:
                            wi_row.append("")
                    
                    table_data.append(wi_row)
                    table_style.append(('FONTSIZE', (6, row_index), (-1, row_index), 6))  # Smaller font for RoC values
                    row_index += 1
                
                # Cost code subtotal
                cc_subtotal_row = [
                    Paragraph("Cost Code Subtotal", normal_style),
                    "",
                    "",
                    Paragraph(f"{cost_code_data['subtotals']['budgeted_quantity']:.2f}", normal_style),
                    Paragraph(f"{cost_code_data['subtotals']['earned_quantity']:.2f}", normal_style),
                    Paragraph(f"{cost_code_data['subtotals']['percent_complete_quantity']:.1f}%", normal_style),
                    "", "", "", "", "", "", ""
                ]
                table_data.append(cc_subtotal_row)
                table_style.append(('SPAN', (0, row_index), (2, row_index)))
                table_style.append(('SPAN', (6, row_index), (-1, row_index)))
                table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
                row_index += 1
            
            # Discipline subtotal
            disc_subtotal_row = [
                Paragraph("Discipline Subtotal", normal_style),
                "",
                "",
                Paragraph(f"{discipline_data['subtotals']['budgeted_quantity']:.2f}", normal_style),
                Paragraph(f"{discipline_data['subtotals']['earned_quantity']:.2f}", normal_style),
                Paragraph(f"{discipline_data['subtotals']['percent_complete_quantity']:.1f}%", normal_style),
                "", "", "", "", "", "", ""
            ]
            table_data.append(disc_subtotal_row)
            table_style.append(('SPAN', (0, row_index), (2, row_index)))
            table_style.append(('SPAN', (6, row_index), (-1, row_index)))
            table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
            row_index += 1
        
        # Total row
        total_row = [
            Paragraph("Total", normal_style),
            "",
            "",
            Paragraph(f"{data['project_totals']['budgeted_quantity']:.2f}", normal_style),
            Paragraph(f"{data['project_totals']['earned_quantity']:.2f}", normal_style),
            Paragraph(f"{data['project_totals']['percent_complete_quantity']:.1f}%", normal_style),
            "", "", "", "", "", "", ""
        ]
        table_data.append(total_row)
        table_style.append(('SPAN', (0, row_index), (2, row_index)))
        table_style.append(('SPAN', (6, row_index), (-1, row_index)))
        table_style.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightblue))
        
        # Create table with column widths
        col_widths = [
            0.7*inch,  # Work Item (10%)
            1.4*inch,  # Description (20%)
            0.5*inch,  # UOM
            0.7*inch,  # Budgeted Qty
            0.7*inch,  # Earned Qty
            0.7*inch,  # % Complete
            0.55*inch, # RoC 1
            0.55*inch, # RoC 2
            0.55*inch, # RoC 3
            0.55*inch, # RoC 4
            0.55*inch, # RoC 5
            0.55*inch, # RoC 6
            0.55*inch  # RoC 7
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return output_path
    
    def generate_excel_export(self, project_id, output_path, sub_job_id=None):
        """Generate Excel export using openpyxl"""
        data = self._get_report_data(project_id, sub_job_id)
        
        # Create workbook and sheets
        wb = openpyxl.Workbook()
        
        # Create Hours sheet
        hours_sheet = wb.active
        hours_sheet.title = "Hours"
        
        # Create Quantities sheet
        quantities_sheet = wb.create_sheet("Quantities")
        
        # Create Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        
        # Styles
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        small_font = Font(size=8)
        
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        discipline_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        subtotal_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        total_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # --- Hours Sheet ---
        # Headers
        hours_sheet.append(["Project:", data['project'].name, data['project'].project_id_str])
        if data['sub_job']:
            hours_sheet.append(["Sub-Job:", data['sub_job'].name, data['sub_job'].sub_job_id_str])
        hours_sheet.append(["Report Date:", data['report_date']])
        hours_sheet.append([])  # Empty row
        
        # Column headers
        headers = ["Work Item", "Description", "Discipline", "Cost Code", "Budgeted Hours", "Earned Hours", "% Complete"]
        
        # Add RoC step headers (up to 7)
        roc_headers = data['all_roc_step_names'][:7]
        while len(roc_headers) < 7:
            roc_headers.append("")
        
        headers.extend(roc_headers)
        hours_sheet.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = hours_sheet.cell(row=5, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Set column widths
        hours_sheet.column_dimensions['A'].width = 15  # Work Item
        hours_sheet.column_dimensions['B'].width = 30  # Description
        hours_sheet.column_dimensions['C'].width = 15  # Discipline
        hours_sheet.column_dimensions['D'].width = 15  # Cost Code
        hours_sheet.column_dimensions['E'].width = 15  # Budgeted Hours
        hours_sheet.column_dimensions['F'].width = 15  # Earned Hours
        hours_sheet.column_dimensions['G'].width = 15  # % Complete
        
        # RoC columns
        for i in range(7):
            col_letter = get_column_letter(8 + i)
            hours_sheet.column_dimensions[col_letter].width = 12
        
        # Add data rows
        row_index = 6
        for discipline, discipline_data in data['grouped_work_items'].items():
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                for work_item in cost_code_data['work_items']:
                    row_data = [
                        work_item.work_item_id_str,
                        work_item.description,
                        discipline,
                        cost_code,
                        work_item.budgeted_man_hours or 0,
                        work_item.earned_man_hours or 0,
                        work_item.percent_complete_hours or 0
                    ]
                    
                    # Add RoC step progress
                    for step_name in roc_headers:
                        if step_name and step_name in work_item.roc_steps_progress:
                            row_data.append(work_item.roc_steps_progress[step_name])
                        else:
                            row_data.append(0)
                    
                    hours_sheet.append(row_data)
                    
                    # Style row
                    for col in range(1, len(row_data) + 1):
                        cell = hours_sheet.cell(row=row_index, column=col)
                        cell.font = normal_font
                        cell.border = border
                        
                        if col in [1, 2, 3, 4]:  # Text columns
                            cell.alignment = left_align
                        else:  # Numeric columns
                            cell.alignment = right_align
                            if col in [5, 6]:  # Hours
                                cell.number_format = '0.00'
                            elif col == 7:  # Percentage
                                cell.number_format = '0.0"%"'
                            elif col >= 8:  # RoC percentages
                                cell.number_format = '0.0"%"'
                                cell.font = small_font
                    
                    row_index += 1
        
        # --- Quantities Sheet ---
        # Headers
        quantities_sheet.append(["Project:", data['project'].name, data['project'].project_id_str])
        if data['sub_job']:
            quantities_sheet.append(["Sub-Job:", data['sub_job'].name, data['sub_job'].sub_job_id_str])
        quantities_sheet.append(["Report Date:", data['report_date']])
        quantities_sheet.append([])  # Empty row
        
        # Column headers
        headers = ["Work Item", "Description", "Discipline", "Cost Code", "UOM", "Budgeted Qty", "Earned Qty", "% Complete"]
        
        # Add RoC step headers (up to 7)
        headers.extend(roc_headers)
        quantities_sheet.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = quantities_sheet.cell(row=5, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Set column widths
        quantities_sheet.column_dimensions['A'].width = 15  # Work Item
        quantities_sheet.column_dimensions['B'].width = 30  # Description
        quantities_sheet.column_dimensions['C'].width = 15  # Discipline
        quantities_sheet.column_dimensions['D'].width = 15  # Cost Code
        quantities_sheet.column_dimensions['E'].width = 10  # UOM
        quantities_sheet.column_dimensions['F'].width = 15  # Budgeted Qty
        quantities_sheet.column_dimensions['G'].width = 15  # Earned Qty
        quantities_sheet.column_dimensions['H'].width = 15  # % Complete
        
        # RoC columns
        for i in range(7):
            col_letter = get_column_letter(9 + i)
            quantities_sheet.column_dimensions[col_letter].width = 12
        
        # Add data rows
        row_index = 6
        for discipline, discipline_data in data['grouped_work_items'].items():
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                for work_item in cost_code_data['work_items']:
                    row_data = [
                        work_item.work_item_id_str,
                        work_item.description,
                        discipline,
                        cost_code,
                        work_item.unit_of_measure or "",
                        work_item.budgeted_quantity or 0,
                        work_item.earned_quantity or 0,
                        work_item.percent_complete_quantity or 0
                    ]
                    
                    # Add RoC step progress
                    for step_name in roc_headers:
                        if step_name and step_name in work_item.roc_steps_progress:
                            row_data.append(work_item.roc_steps_progress[step_name])
                        else:
                            row_data.append(0)
                    
                    quantities_sheet.append(row_data)
                    
                    # Style row
                    for col in range(1, len(row_data) + 1):
                        cell = quantities_sheet.cell(row=row_index, column=col)
                        cell.font = normal_font
                        cell.border = border
                        
                        if col in [1, 2, 3, 4, 5]:  # Text columns
                            cell.alignment = left_align
                        else:  # Numeric columns
                            cell.alignment = right_align
                            if col in [6, 7]:  # Quantities
                                cell.number_format = '0.00'
                            elif col == 8:  # Percentage
                                cell.number_format = '0.0"%"'
                            elif col >= 9:  # RoC percentages
                                cell.number_format = '0.0"%"'
                                cell.font = small_font
                    
                    row_index += 1
        
        # --- Summary Sheet ---
        # Headers
        summary_sheet.append(["Project:", data['project'].name, data['project'].project_id_str])
        if data['sub_job']:
            summary_sheet.append(["Sub-Job:", data['sub_job'].name, data['sub_job'].sub_job_id_str])
        summary_sheet.append(["Report Date:", data['report_date']])
        summary_sheet.append([])  # Empty row
        
        # Hours summary
        summary_sheet.append(["Hours Summary"])
        summary_sheet.append(["Discipline", "Cost Code", "Budgeted Hours", "Earned Hours", "% Complete"])
        
        # Style headers
        for col in range(1, 6):
            cell = summary_sheet.cell(row=6, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Add discipline and cost code summaries
        row_index = 7
        for discipline, discipline_data in data['grouped_work_items'].items():
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                row_data = [
                    discipline,
                    cost_code,
                    cost_code_data['subtotals']['budgeted_hours'],
                    cost_code_data['subtotals']['earned_hours'],
                    cost_code_data['subtotals']['percent_complete_hours']
                ]
                
                summary_sheet.append(row_data)
                
                # Style row
                for col in range(1, 6):
                    cell = summary_sheet.cell(row=row_index, column=col)
                    cell.font = normal_font
                    cell.border = border
                    
                    if col in [1, 2]:  # Text columns
                        cell.alignment = left_align
                    else:  # Numeric columns
                        cell.alignment = right_align
                        if col in [3, 4]:  # Hours
                            cell.number_format = '0.00'
                        elif col == 5:  # Percentage
                            cell.number_format = '0.0"%"'
                
                row_index += 1
            
            # Discipline subtotal
            row_data = [
                f"{discipline} Subtotal",
                "",
                discipline_data['subtotals']['budgeted_hours'],
                discipline_data['subtotals']['earned_hours'],
                discipline_data['subtotals']['percent_complete_hours']
            ]
            
            summary_sheet.append(row_data)
            
            # Style subtotal row
            for col in range(1, 6):
                cell = summary_sheet.cell(row=row_index, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = subtotal_fill
                cell.border = border
                
                if col == 1:  # Text column
                    cell.alignment = left_align
                else:  # Numeric columns
                    cell.alignment = right_align
                    if col in [3, 4]:  # Hours
                        cell.number_format = '0.00'
                    elif col == 5:  # Percentage
                        cell.number_format = '0.0"%"'
            
            # Merge discipline subtotal cells
            summary_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
            
            row_index += 1
        
        # Project total
        row_data = [
            "Project Total",
            "",
            data['project_totals']['budgeted_hours'],
            data['project_totals']['earned_hours'],
            data['project_totals']['percent_complete_hours']
        ]
        
        summary_sheet.append(row_data)
        
        # Style total row
        for col in range(1, 6):
            cell = summary_sheet.cell(row=row_index, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = total_fill
            cell.border = border
            
            if col == 1:  # Text column
                cell.alignment = left_align
            else:  # Numeric columns
                cell.alignment = right_align
                if col in [3, 4]:  # Hours
                    cell.number_format = '0.00'
                elif col == 5:  # Percentage
                    cell.number_format = '0.0"%"'
        
        # Merge total cells
        summary_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
        
        row_index += 2
        
        # Quantities summary
        summary_sheet.append(["Quantities Summary"])
        summary_sheet.append(["Discipline", "Cost Code", "Budgeted Qty", "Earned Qty", "% Complete"])
        
        # Style headers
        for col in range(1, 6):
            cell = summary_sheet.cell(row=row_index, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        row_index += 1
        
        # Add discipline and cost code summaries
        for discipline, discipline_data in data['grouped_work_items'].items():
            for cost_code, cost_code_data in discipline_data['cost_codes'].items():
                row_data = [
                    discipline,
                    cost_code,
                    cost_code_data['subtotals']['budgeted_quantity'],
                    cost_code_data['subtotals']['earned_quantity'],
                    cost_code_data['subtotals']['percent_complete_quantity']
                ]
                
                summary_sheet.append(row_data)
                
                # Style row
                for col in range(1, 6):
                    cell = summary_sheet.cell(row=row_index, column=col)
                    cell.font = normal_font
                    cell.border = border
                    
                    if col in [1, 2]:  # Text columns
                        cell.alignment = left_align
                    else:  # Numeric columns
                        cell.alignment = right_align
                        if col in [3, 4]:  # Quantities
                            cell.number_format = '0.00'
                        elif col == 5:  # Percentage
                            cell.number_format = '0.0"%"'
                
                row_index += 1
            
            # Discipline subtotal
            row_data = [
                f"{discipline} Subtotal",
                "",
                discipline_data['subtotals']['budgeted_quantity'],
                discipline_data['subtotals']['earned_quantity'],
                discipline_data['subtotals']['percent_complete_quantity']
            ]
            
            summary_sheet.append(row_data)
            
            # Style subtotal row
            for col in range(1, 6):
                cell = summary_sheet.cell(row=row_index, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = subtotal_fill
                cell.border = border
                
                if col == 1:  # Text column
                    cell.alignment = left_align
                else:  # Numeric columns
                    cell.alignment = right_align
                    if col in [3, 4]:  # Quantities
                        cell.number_format = '0.00'
                    elif col == 5:  # Percentage
                        cell.number_format = '0.0"%"'
            
            # Merge discipline subtotal cells
            summary_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
            
            row_index += 1
        
        # Project total
        row_data = [
            "Project Total",
            "",
            data['project_totals']['budgeted_quantity'],
            data['project_totals']['earned_quantity'],
            data['project_totals']['percent_complete_quantity']
        ]
        
        summary_sheet.append(row_data)
        
        # Style total row
        for col in range(1, 6):
            cell = summary_sheet.cell(row=row_index, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = total_fill
            cell.border = border
            
            if col == 1:  # Text column
                cell.alignment = left_align
            else:  # Numeric columns
                cell.alignment = right_align
                if col in [3, 4]:  # Quantities
                    cell.number_format = '0.00'
                elif col == 5:  # Percentage
                    cell.number_format = '0.0"%"'
        
        # Merge total cells
        summary_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def generate_reports(self, project_id, output_dir, sub_job_id=None):
        """Generate all reports for a project or sub-job"""
        os.makedirs(output_dir, exist_ok=True)
        
        # Get project and sub-job info
        cursor = self.conn.cursor()
        cursor.execute("SELECT project_id_str, name FROM project WHERE id = ?", (project_id,))
        project_row = cursor.fetchone()
        
        if not project_row:
            raise ValueError(f"Project with ID {project_id} not found")
        
        project_id_str = project_row['project_id_str']
        project_name = project_row['name']
        
        sub_job_id_str = ""
        sub_job_name = ""
        
        if sub_job_id:
            cursor.execute("SELECT sub_job_id_str, name FROM sub_job WHERE id = ?", (sub_job_id,))
            sub_job_row = cursor.fetchone()
            
            if not sub_job_row:
                raise ValueError(f"Sub-job with ID {sub_job_id} not found")
            
            sub_job_id_str = sub_job_row['sub_job_id_str']
            sub_job_name = sub_job_row['name']
        
        # Generate timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate file names
        if sub_job_id:
            base_name = f"{project_id_str}_{sub_job_id_str}"
            display_name = f"{project_name}_{sub_job_name}"
        else:
            base_name = project_id_str
            display_name = project_name
        
        hours_pdf_path = os.path.join(output_dir, f"{base_name}_hours_{timestamp}.pdf")
        quantities_pdf_path = os.path.join(output_dir, f"{base_name}_quantities_{timestamp}.pdf")
        excel_path = os.path.join(output_dir, f"{base_name}_data_{timestamp}.xlsx")
        
        # Generate reports
        self.generate_hours_pdf(project_id, hours_pdf_path, sub_job_id)
        self.generate_quantities_pdf(project_id, quantities_pdf_path, sub_job_id)
        self.generate_excel_export(project_id, excel_path, sub_job_id)
        
        return {
            'hours_pdf': hours_pdf_path,
            'quantities_pdf': quantities_pdf_path,
            'excel': excel_path
        }
