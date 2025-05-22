import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models import Project, SubJob, WorkItem, CostCode, RuleOfCredit
from reports.pdf_export import prepare_report_data

def generate_excel_report(data, output_path, report_type):
    """Generate Excel report from data"""
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set the title based on report type
    if report_type == 'hours':
        ws.title = "Hours Report"
        report_title = "PROGRESS DETAIL"
        report_subtitle = "Sub Job by Hours" if data['sub_job'] else "Total Project by Hours"
    else:
        ws.title = "Quantities Report"
        report_title = "QUANTITY DETAIL"
        report_subtitle = "Sub Job by Quantities" if data['sub_job'] else "Total Project by Quantities"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=8)
    small_font = Font(name='Arial', size=6)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    discipline_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    cost_code_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    subtotal_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
    total_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create header section
    ws.merge_cells('A1:C3')
    ws.merge_cells('D1:G3')
    ws.merge_cells('H1:J3')
    
    # Logo placeholder (left section)
    ws['A1'] = "MAGELLAN"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    
    # Center section with report title and project info
    ws['D1'] = f"{report_title}\n{report_subtitle}\n{data['project'].name}\n{data['project'].project_id_str}"
    ws['D1'].font = header_font
    ws['D1'].alignment = center_align
    
    # Right section with date
    ws['H1'] = f"Page 1\nProgress Thru:\n{data['report_date']}"
    ws['H1'].font = subheader_font
    ws['H1'].alignment = right_align
    
    # Add sub job info if applicable
    current_row = 4
    if data['sub_job']:
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"{data['sub_job'].sub_job_id_str} - {data['sub_job'].name}"
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].alignment = left_align
        current_row += 1
    
    # Create table headers
    start_col = 1  # Column A
    if report_type == 'hours':
        headers = ['Work Item', 'Description', 'Budgeted Hours', 'Earned Hours', '% Complete', 'Rules of Credit Steps']
        col_widths = [10, 30, 12, 12, 10, 10, 10, 10, 10, 10, 10]  # Width for each column
    else:
        headers = ['Work Item', 'Description', 'UOM', 'Budgeted Quantity', 'Earned Quantity', '% Complete', 'Rules of Credit Steps']
        col_widths = [10, 30, 8, 15, 15, 10, 10, 10, 10, 10, 10, 10]  # Width for each column
    
    # Set column widths
    for i, width in enumerate(col_widths):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = width
    
    # Write headers
    for i, header in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = normal_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        
        # Merge Rules of Credit Steps across 7 columns
        if header == 'Rules of Credit Steps':
            end_col = col + 6
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=end_col)
    
    current_row += 1
    
    # Process grouped work items
    for discipline, discipline_data in data['grouped_work_items'].items():
        # Add discipline header
        discipline_col_span = 12 if report_type == 'quantities' else 11
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=discipline_col_span)
        cell = ws.cell(row=current_row, column=1)
        cell.value = discipline
        cell.font = normal_font
        cell.alignment = left_align
        cell.fill = discipline_fill
        cell.border = thin_border
        current_row += 1
        
        # Process cost codes
        for cost_code, cost_code_data in discipline_data['cost_codes'].items():
            # Add cost code header
            cell = ws.cell(row=current_row, column=1)
            cell.value = cost_code
            cell.font = normal_font
            cell.alignment = left_align
            cell.fill = cost_code_fill
            cell.border = thin_border
            
            # Merge empty cells for other columns
            if report_type == 'hours':
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                empty_cell = ws.cell(row=current_row, column=2)
                empty_cell.fill = cost_code_fill
                empty_cell.border = thin_border
            else:
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
                empty_cell = ws.cell(row=current_row, column=2)
                empty_cell.fill = cost_code_fill
                empty_cell.border = thin_border
            
            # Add ROC step headers
            roc_col_start = 6 if report_type == 'hours' else 7
            for i, step_header in enumerate(cost_code_data['roc_step_headers']):
                if step_header:  # Only add non-empty headers
                    cell = ws.cell(row=current_row, column=roc_col_start + i)
                    cell.value = step_header
                    cell.font = small_font
                    cell.alignment = center_align
                    cell.fill = cost_code_fill
                    cell.border = thin_border
            
            current_row += 1
            
            # Add work items
            for work_item in cost_code_data['work_items']:
                # Work Item ID
                cell = ws.cell(row=current_row, column=1)
                cell.value = work_item['work_item_id_str']
                cell.font = normal_font
                cell.alignment = center_align
                cell.border = thin_border
                
                # Description
                cell = ws.cell(row=current_row, column=2)
                cell.value = work_item['description']
                cell.font = normal_font
                cell.alignment = left_align
                cell.border = thin_border
                
                if report_type == 'hours':
                    # Budgeted Hours
                    cell = ws.cell(row=current_row, column=3)
                    cell.value = work_item['budgeted_man_hours']
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.00'
                    
                    # Earned Hours
                    cell = ws.cell(row=current_row, column=4)
                    cell.value = work_item['earned_man_hours']
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.00'
                    
                    # % Complete
                    cell = ws.cell(row=current_row, column=5)
                    cell.value = work_item['percent_complete_hours'] / 100  # Convert to decimal for Excel percentage
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.0%'
                    
                    # ROC Steps Progress
                    for i, step_name in enumerate(cost_code_data['roc_step_headers']):
                        if step_name and step_name in work_item['roc_steps_progress']:
                            cell = ws.cell(row=current_row, column=6 + i)
                            cell.value = work_item['roc_steps_progress'][step_name] / 100  # Convert to decimal
                            cell.font = small_font
                            cell.alignment = right_align
                            cell.border = thin_border
                            cell.number_format = '0.0%'
                else:
                    # UOM
                    cell = ws.cell(row=current_row, column=3)
                    cell.value = work_item['unit_of_measure']
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Budgeted Quantity
                    cell = ws.cell(row=current_row, column=4)
                    cell.value = work_item['budgeted_quantity']
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.00'
                    
                    # Earned Quantity
                    cell = ws.cell(row=current_row, column=5)
                    cell.value = work_item['earned_quantity']
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.00'
                    
                    # % Complete
                    cell = ws.cell(row=current_row, column=6)
                    cell.value = work_item['percent_complete_quantity'] / 100  # Convert to decimal for Excel percentage
                    cell.font = normal_font
                    cell.alignment = right_align
                    cell.border = thin_border
                    cell.number_format = '0.0%'
                    
                    # ROC Steps Progress
                    for i, step_name in enumerate(cost_code_data['roc_step_headers']):
                        if step_name and step_name in work_item['roc_steps_progress']:
                            cell = ws.cell(row=current_row, column=7 + i)
                            cell.value = work_item['roc_steps_progress'][step_name] / 100  # Convert to decimal
                            cell.font = small_font
                            cell.alignment = right_align
                            cell.border = thin_border
                            cell.number_format = '0.0%'
                
                current_row += 1
            
            # Add cost code subtotal
            if report_type == 'hours':
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell = ws.cell(row=current_row, column=1)
                cell.value = "Cost Code Subtotal"
                cell.font = normal_font
                cell.alignment = left_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                
                # Budgeted Hours
                cell = ws.cell(row=current_row, column=3)
                cell.value = cost_code_data['subtotals']['budgeted_hours']
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.00'
                
                # Earned Hours
                cell = ws.cell(row=current_row, column=4)
                cell.value = cost_code_data['subtotals']['earned_hours']
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.00'
                
                # % Complete
                cell = ws.cell(row=current_row, column=5)
                cell.value = cost_code_data['subtotals']['percent_complete_hours'] / 100
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.0%'
                
                # Empty cells for ROC steps
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=12)
                empty_cell = ws.cell(row=current_row, column=6)
                empty_cell.fill = subtotal_fill
                empty_cell.border = thin_border
            else:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell = ws.cell(row=current_row, column=1)
                cell.value = "Cost Code Subtotal"
                cell.font = normal_font
                cell.alignment = left_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                
                # Empty UOM cell
                cell = ws.cell(row=current_row, column=3)
                cell.fill = subtotal_fill
                cell.border = thin_border
                
                # Budgeted Quantity
                cell = ws.cell(row=current_row, column=4)
                cell.value = cost_code_data['subtotals']['budgeted_quantity']
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.00'
                
                # Earned Quantity
                cell = ws.cell(row=current_row, column=5)
                cell.value = cost_code_data['subtotals']['earned_quantity']
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.00'
                
                # % Complete
                cell = ws.cell(row=current_row, column=6)
                cell.value = cost_code_data['subtotals']['percent_complete_quantity'] / 100
                cell.font = normal_font
                cell.alignment = right_align
                cell.fill = subtotal_fill
                cell.border = thin_border
                cell.number_format = '0.0%'
                
                # Empty cells for ROC steps
                ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=13)
                empty_cell = ws.cell(row=current_row, column=7)
                empty_cell.fill = subtotal_fill
                empty_cell.border = thin_border
            
            current_row += 1
        
        # Add discipline subtotal
        if report_type == 'hours':
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell = ws.cell(row=current_row, column=1)
            cell.value = "Discipline Subtotal"
            cell.font = normal_font
            cell.alignment = left_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            
            # Budgeted Hours
            cell = ws.cell(row=current_row, column=3)
            cell.value = discipline_data['subtotals']['budgeted_hours']
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.00'
            
            # Earned Hours
            cell = ws.cell(row=current_row, column=4)
            cell.value = discipline_data['subtotals']['earned_hours']
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.00'
            
            # % Complete
            cell = ws.cell(row=current_row, column=5)
            cell.value = discipline_data['subtotals']['percent_complete_hours'] / 100
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.0%'
            
            # Empty cells for ROC steps
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=12)
            empty_cell = ws.cell(row=current_row, column=6)
            empty_cell.fill = subtotal_fill
            empty_cell.border = thin_border
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            cell = ws.cell(row=current_row, column=1)
            cell.value = "Discipline Subtotal"
            cell.font = normal_font
            cell.alignment = left_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            
            # Budgeted Quantity
            cell = ws.cell(row=current_row, column=4)
            cell.value = discipline_data['subtotals']['budgeted_quantity']
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.00'
            
            # Earned Quantity
            cell = ws.cell(row=current_row, column=5)
            cell.value = discipline_data['subtotals']['earned_quantity']
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.00'
            
            # % Complete
            cell = ws.cell(row=current_row, column=6)
            cell.value = discipline_data['subtotals']['percent_complete_quantity'] / 100
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = subtotal_fill
            cell.border = thin_border
            cell.number_format = '0.0%'
            
            # Empty cells for ROC steps
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=13)
            empty_cell = ws.cell(row=current_row, column=7)
            empty_cell.fill = subtotal_fill
            empty_cell.border = thin_border
        
        current_row += 1
    
    # Add project total
    if report_type == 'hours':
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Total"
        cell.font = normal_font
        cell.alignment = left_align
        cell.fill = total_fill
        cell.border = thin_border
        
        # Budgeted Hours
        cell = ws.cell(row=current_row, column=3)
        cell.value = data['project_totals']['budgeted_hours']
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # Earned Hours
        cell = ws.cell(row=current_row, column=4)
        cell.value = data['project_totals']['earned_hours']
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # % Complete
        cell = ws.cell(row=current_row, column=5)
        cell.value = data['project_totals']['percent_complete_hours'] / 100
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.0%'
        
        # Empty cells for ROC steps
        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=12)
        empty_cell = ws.cell(row=current_row, column=6)
        empty_cell.fill = total_fill
        empty_cell.border = thin_border
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Total"
        cell.font = normal_font
        cell.alignment = left_align
        cell.fill = total_fill
        cell.border = thin_border
        
        # Budgeted Quantity
        cell = ws.cell(row=current_row, column=4)
        cell.value = data['project_totals']['budgeted_quantity']
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # Earned Quantity
        cell = ws.cell(row=current_row, column=5)
        cell.value = data['project_totals']['earned_quantity']
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # % Complete
        cell = ws.cell(row=current_row, column=6)
        cell.value = data['project_totals']['percent_complete_quantity'] / 100
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '0.0%'
        
        # Empty cells for ROC steps
        ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=13)
        empty_cell = ws.cell(row=current_row, column=7)
        empty_cell.fill = total_fill
        empty_cell.border = thin_border
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def generate_hours_report_excel(project_id, sub_job_id=None, output_path=None):
    """Generate hours report Excel"""
    if output_path is None:
        project = Project.query.get_or_404(project_id)
        filename = f"hours_report_{project.project_id_str}"
        if sub_job_id:
            sub_job = SubJob.query.get_or_404(sub_job_id)
            filename += f"_{sub_job.sub_job_id_str}"
        filename += f"_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    data = prepare_report_data(project_id, sub_job_id, 'hours')
    return generate_excel_report(data, output_path, 'hours')

def generate_quantities_report_excel(project_id, sub_job_id=None, output_path=None):
    """Generate quantities report Excel"""
    if output_path is None:
        project = Project.query.get_or_404(project_id)
        filename = f"quantities_report_{project.project_id_str}"
        if sub_job_id:
            sub_job = SubJob.query.get_or_404(sub_job_id)
            filename += f"_{sub_job.sub_job_id_str}"
        filename += f"_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    data = prepare_report_data(project_id, sub_job_id, 'quantities')
    return generate_excel_report(data, output_path, 'quantities')
